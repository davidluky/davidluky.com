#!/usr/bin/env python3
"""Regenerate src/data/gameTracker.json from David's "Video Jogos" workbook.

Usage:
    python scripts/generate-game-tracker.py "<path to the current Video Jogos workbook>.xlsx"

Dev-machine tool: the JSON it writes is committed; nothing at build/deploy
time runs this. Needs openpyxl. Reads cached formula values, so the workbook
must have been saved by Excel (not merely edited by a script).

Sheet columns are read BY HEADER NAME (row 4 of every list sheet), so inserting
or reordering columns is safe; renaming or removing one fails loudly with the
headers it actually found.
"""
import datetime as dt
import json
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "src" / "data" / "gameTracker.json"
HEADER_ROW = 4

QUEUE_STATUS = {
    "Na fila": "Backlog",
    "Meta 2026": "2026 target",
    "Jogando agora": "In progress",
    "JA TERMINADO": None,  # finished games stay only in history
}
PLAYED_STATUS = {"Terminado": "Finished", "Em andamento": "In progress"}
# Same PT source values as the queue, but this sheet keeps finished entries visible.
POKEMON_STATUS = {**QUEUE_STATUS, "JA TERMINADO": "Finished"}

QUERO_JOGAR_HEADERS = [
    "Jogo", "Versao / plataforma", "Lancamento", "Ano", "Decada", "Onde jogar",
    "Trilha", "Status", "Fila #", "Prioridade", "Notas", "Fonte da data", "Poke #",
]
TERMINAR_HEADERS = [
    "Jogo", "Versao / observacao", "Lancamento", "Onde jogar", "Comecei em",
    "Status", "Terminar #", "Notas",
]
JOGADOS_HEADERS = [
    "Jogo", "Inicio", "Fim", "Status", "Dias", "Ano de conclusao",
    "Fonte original", "Observacao", "Terminei sem lembrar a data?", "Terminar #",
]
POKEMON_SHEET_HEADERS = [
    "#", "Jogo (versoes separadas por / )", "Plataforma", "Lancamento",
    "Onde jogar", "Status",
]
POKEMON_NAME = "Jogo (versoes separadas por / )"


def clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def iso(value):
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date) and value.year > 1900:
        return value.isoformat()
    return None  # times, strings, blanks, 1899/1900 Excel blank-cell sentinels


def year_of(iso_date):
    return int(iso_date[:4]) if iso_date else None


def as_int(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    return None


def norm_name(value):
    text = unicodedata.normalize("NFD", str(value)).casefold()
    return "".join(ch for ch in text if not unicodedata.combining(ch)).strip()


def sheet_rows(ws, expected_headers, key_column):
    """Yield the sheet's data rows as dicts keyed by header name.

    Every name in expected_headers must exist in the header row; positional
    reads are never used, so a column insert or reorder cannot shift the data.
    """
    header = [clean(cell) for cell in next(
        ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)
    )]
    index = {name: position for position, name in enumerate(header) if name is not None}
    missing = [name for name in expected_headers if name not in index]
    if missing:
        raise SystemExit(
            f"{ws.title}: missing expected header(s) {missing} in row {HEADER_ROW}. "
            f"Headers found: {[name for name in header if name is not None]}"
        )

    rows = []
    for raw in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        row = {name: (raw[position] if position < len(raw) else None)
               for name, position in index.items()}
        if clean(row[key_column]) is None:
            continue  # trailing/blank rows, and formula rows that only carry a number
        rows.append(row)
    return rows


def map_status(raw, table, sheet):
    status = clean(raw)
    if status not in table:
        raise SystemExit(f"{sheet}: unknown status {status!r}")
    return table[status]


def join_notes(*parts):
    text = " · ".join(p for p in (clean(part) for part in parts) if p)
    return text or None


def main():
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    src = Path(sys.argv[1])
    wb = openpyxl.load_workbook(src, data_only=True)

    # --- Quero jogar -> playOrder ---
    play_order = []
    for row in sheet_rows(wb["Quero jogar"], QUERO_JOGAR_HEADERS, "Jogo"):
        status = map_status(row["Status"], QUEUE_STATUS, "Quero jogar")
        if status is None:
            continue
        game = clean(row["Jogo"])
        decade = clean(row["Decada"])
        if decade is None:
            raise SystemExit(f"Quero jogar: missing 'Decada' for {game!r}")
        release = iso(row["Lancamento"])
        play_order.append({
            "rank": len(play_order) + 1,
            "game": game,
            "notes": join_notes(row["Versao / plataforma"], row["Notas"]),
            "releaseDate": release,
            "year": as_int(row["Ano"]) or year_of(release),
            "decade": decade,
            "whereToPlay": clean(row["Onde jogar"]),
            "trilha": clean(row["Trilha"]),
            "status": status,
            "queueRank": as_int(row["Fila #"]),
            "dateSource": clean(row["Fonte da data"]),
            "pokeNumber": as_int(row["Poke #"]),
        })

    # --- Terminar este ano -> finishThisYear (Pendente only) ---
    finish = []
    for row in sheet_rows(wb["Terminar este ano"], TERMINAR_HEADERS, "Jogo"):
        if clean(row["Status"]) != "Pendente":
            continue
        finish.append({
            "priority": as_int(row["Terminar #"]),
            "game": clean(row["Jogo"]),
            "notes": join_notes(row["Versao / observacao"], row["Notas"]),
            "releaseDate": iso(row["Lancamento"]),
            "whereToPlay": clean(row["Onde jogar"]),
            "started": iso(row["Comecei em"]),
        })
    if sorted(r["priority"] for r in finish) != list(range(1, len(finish) + 1)):
        raise SystemExit(f"Terminar este ano: priorities not 1..N: {[r['priority'] for r in finish]}")
    finish.sort(key=lambda r: r["priority"])

    # --- Jogados -> history / inProgress / finishedThisYear ---
    # In-progress rows borrow release date and platform from whichever sheet
    # knows the game: the queue first, then this year's finish list.
    queue_by_name = {norm_name(r["game"]): r for r in play_order}
    finish_by_name = {norm_name(r["game"]): r for r in finish}
    history, in_progress, finished_2026 = [], [], []
    for row in sheet_rows(wb["Jogados"], JOGADOS_HEADERS, "Jogo"):
        game = clean(row["Jogo"])
        status = map_status(row["Status"], PLAYED_STATUS, "Jogados")
        started, finished = iso(row["Inicio"]), iso(row["Fim"])
        stated_year = as_int(row["Ano de conclusao"])
        if status == "Finished":
            year = stated_year or year_of(finished) or year_of(started)
        else:
            year = stated_year or year_of(started) or year_of(finished)
        if year is None:
            raise SystemExit(
                f"Jogados: cannot determine a year for {game!r} "
                "('Ano de conclusao', 'Fim' and 'Inicio' are all empty)"
            )
        record = {
            "year": year,
            "game": game,
            "started": started,
            "finished": finished,
            "status": status,
            "days": as_int(row["Dias"]),
            "source": clean(row["Fonte original"]),
            "notes": clean(row["Observacao"]),
        }
        history.append(record)
        if status == "In progress":
            key = norm_name(game)
            known = queue_by_name.get(key) or finish_by_name.get(key)
            if known is None:
                print(
                    "WARNING: in-progress game not found in Quero jogar or "
                    f"Terminar este ano: {game}"
                )
            in_progress.append({
                "game": game,
                "started": started,
                "releaseDate": known["releaseDate"] if known else None,
                "whereToPlay": known["whereToPlay"] if known else None,
                "notes": record["notes"],
                "source": record["source"],
            })
        elif year == 2026:
            finished_2026.append(record)
    in_progress.sort(key=lambda r: r["started"] or "9999")
    finished_2026.sort(key=lambda r: r["finished"] or "9999")

    # --- Pokemon sheet ---
    # Trailing rows keep their formula-driven number but have no game name,
    # so the row filter keys on the name column.
    pokemon = []
    for row in sheet_rows(wb["Pokemon"], POKEMON_SHEET_HEADERS, POKEMON_NAME):
        rank = as_int(row["#"])
        if rank is None:
            continue
        pokemon.append({
            "rank": rank,
            "game": clean(row[POKEMON_NAME]),
            "platform": clean(row["Plataforma"]),
            "releaseDate": iso(row["Lancamento"]),
            "whereToPlay": clean(row["Onde jogar"]),
            "status": map_status(row["Status"], POKEMON_STATUS, "Pokemon"),
        })

    # --- Painel -> forecast numbers (match by exact stripped label in col A) ---
    painel = {}
    termina_dates = []
    for row in wb["Painel"].iter_rows(values_only=True):
        label, value = clean(row[0]), row[1]
        if label is None:
            continue
        if label == "termina por volta de":
            termina_dates.append(iso(value))
        else:
            painel[label] = value
    try:
        pace = round(float(painel["Ritmo: terminados por mes (media 2022-2025)"]), 2)
        remaining = as_int(painel["Faltam terminar (metas deste ano + fila)"])
        total_mapped = as_int(painel["Total do acervo mapeado"])
        scenario = as_int(painel["E se voce jogar este tanto por mes:  (pode mudar)"])
    except KeyError as missing:
        raise SystemExit(f"Painel: label not found: {missing}. Labels seen: {sorted(painel)}")
    if len(termina_dates) != 2 or None in termina_dates:
        raise SystemExit(f"Painel: expected 2 'termina por volta de' dates, got {termina_dates}")

    decade_counts = dict(sorted(Counter(r["decade"] for r in play_order).items()))
    updated = dt.date.fromtimestamp(src.stat().st_mtime).isoformat()

    data = {
        "updated": updated,
        "generatedFrom": f"{src.name} ({updated})",
        "summary": {
            "activePlayOrderRows": len(play_order),
            "backlogRows": sum(1 for r in play_order if r["status"] == "Backlog"),
            "queue2026Targets": sum(1 for r in play_order if r["status"] == "2026 target"),
            "queueInProgress": sum(1 for r in play_order if r["status"] == "In progress"),
            "finishThisYear": len(finish),
            "inProgress": len(in_progress),
            "finishedThisYear": len(finished_2026),
            "previousFinished": sum(1 for r in history if r["status"] == "Finished" and r["year"] < 2026),
            "pokemonRows": len(pokemon),
            "historyRecords": len(history),
            "oldestActiveGame": play_order[0]["game"],
            "oldestReleaseYear": play_order[0]["year"],
            "nextFinishTarget": finish[0]["game"],
            "totalMapped": total_mapped,
            "remainingToFinish": remaining,
            "paceFinishedPerMonth": pace,
            "scenarioPerMonth": scenario,
            "projectedFinishCurrentPace": termina_dates[0],
            "projectedFinishOnePerMonth": termina_dates[1],
        },
        "playOrder": play_order,
        "finishThisYear": finish,
        "inProgress": in_progress,
        "finishedThisYear": finished_2026,
        "pokemon": pokemon,
        "history": history,
        "decadeCounts": decade_counts,
    }

    with open(OUT_PATH, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(json.dumps(data["summary"], ensure_ascii=False, indent=2))
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
